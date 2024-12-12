import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from bs4 import BeautifulSoup
import time
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
from PIL import Image
from io import BytesIO
import re
from urllib.parse import unquote
import undetected_chromedriver as uc
from selenium.webdriver.chrome.options import Options


USERNAME = ''  # Enter to your actual username
PASSWORD = ''  # Enter to your actual password


def initialize_webdriver(proxy=None):
    options = Options()

    # Set up headless mode (you can set to False if you need the browser UI)
    options.headless = False  # Change to True if you want to run headless


    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")


    options.add_argument('--disable-blink-features=AutomationControlled')


    options.add_argument("enable-automation")
    options.add_argument("--disable-features=IsolateOrigins,site-per-process")


    if proxy:
        options.add_argument(f'--proxy-server={proxy}')


    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')


    options.add_argument('--disable-gpu')


    options.add_argument("--enable-logging")
    options.add_argument("--v=1")

    # Initialize the WebDriver
    driver = uc.Chrome(options=options)

    return driver



# Function to log in to the website
def login(driver):
    driver.get("https://chatgpt.com/auth/login?next=/chat")
    time.sleep(8)

    try:
        initial_login_button = driver.find_element(By.XPATH, '//button[@data-testid="login-button"]')
        initial_login_button.click()
        time.sleep(30)

        email_field = driver.find_element(By.ID, 'email-input')
        email_field.send_keys(USERNAME)
        time.sleep(8)

        login_button = driver.find_element(By.XPATH, '/html/body/div/main/section/div[2]/button')
        login_button.click()
        time.sleep(10)

        password_field = driver.find_element(By.ID, 'password')
        password_field.send_keys(PASSWORD)
        time.sleep(8)

        submit_button = driver.find_element(By.XPATH,
                                            '//*[@id="auth0-widget"]/main/section/div/div/div/form/div[2]/button')
        submit_button.click()
        time.sleep(10)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "profile-icon"))
        )

        print("Successfully logged in!")
    except Exception as e:
        print(f"Login failed: {e}")



# Function to select Excel file
def select_excel_file():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    if file_path:
        return file_path
    else:
        messagebox.showwarning("File Selection", "No file selected!")
        return None


# Function to load URLs from Excel
def load_urls_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    urls = []
    for row in range(2, sheet.max_row + 1):  # Loop through all rows
        for col in range(2, sheet.max_column + 1):  # Loop through columns with URLs
            url = sheet.cell(row=row, column=col).value
            if isinstance(url, str) and url.startswith("https://chatgpt.com"):
                urls.append((url, row))
    return urls, workbook



# Function to create a new sheet in Excel
def create_new_sheet(workbook, sheet_name="ScrapedData"):
    if sheet_name in workbook.sheetnames:
        new_sheet = workbook[sheet_name]
    else:
        new_sheet = workbook.create_sheet(title=sheet_name)
    headers = [
        "Retrieved Date", "Retrieved Time", "URL", "Images", "DALL-E", "Image Creation", "DALL-E Prompt",
        "Title", "Creator", "Creator Website", "Creator LinkedIn", "Creator GitHub", "Creator X",
        "Description", "Sample Message 1", "Sample Message 2", "Sample Message 3", "Prompt"
    ]
    for col_num, header in enumerate(headers, 1):
        new_sheet.cell(row=1, column=col_num, value=header)
    return new_sheet



# Function to extract image data
def extract_image_data(images):
    image_creation = ""
    dall_e_prompt = ""

    for image_url in images:
        try:
            print(f"Processing image URL: {image_url}")  # Debugging: Print the image URL being processed

            # Download the image
            response = requests.get(image_url)
            time.sleep(8)
            response.raise_for_status()  # Check if the request was successful
            time.sleep(8)
            print(f"Image downloaded successfully from {image_url}")  # Debugging: Confirm successful download

            # Open the image directly from memory
            image = Image.open(BytesIO(response.content))
            time.sleep(8)
            print("Image opened successfully")  # Debugging: Confirm the image was opened

            # Create a local filename for the downloaded image
            filename = image_url.split("/")[-1]  # Get last part of the URL as filename
            time.sleep(8)
            print(f"Generated filename: {filename}")  # Debugging: Print the generated filename

            # Save the image locally
            with open(filename, 'wb') as file:
                file.write(response.content)
            print(f"Image saved as {filename}")  # Debugging: Confirm the image was saved locally

            # Extract date from the filename (format: "DALL·E YYYY-MM-DD ...")
            date_match = re.search(r"(\d{4})-(\d{2})-(\d{2})", filename)
            if date_match:
                # Extract the date and reformat it as MM-DD-YYYY
                year, month, day = date_match.groups()
                image_creation = f"{month}-{day}-{year}"  # Format as MM-DD-YYYY
                print(f"Extracted date: {image_creation}")  # Debugging: Print the extracted date
            else:
                print("No date found in filename.")  # Debugging: Print if no date was found

            # Check if the filename contains a prompt (e.g., DALL-E prompt or other text)
            text_match = re.search(r"(\d{4}-\d{2}-\d{2})\s+-\s+(.*)", filename)
            if text_match:
                # Extract the prompt text after the date and clean it
                text = unquote(text_match.group(2))
                dall_e_prompt = text.replace("%2C", ",").replace("%20", " ").replace("%22", "\"")
                print(f"Extracted prompt: {dall_e_prompt}")  # Debugging: Print the extracted prompt
            else:
                print("No prompt found in filename.")  # Debugging: Print if no prompt was found

        except requests.exceptions.RequestException as e:
            print(f"Error fetching image from {image_url}: {e}")
            continue  # Skip this image URL if there's an error
        except Exception as e:
            print(f"Unexpected error with image {image_url}: {e}")
            continue  # Skip this image URL if there's an error

    # Set DALL-E flag based on whether a prompt was found
    dall_e = "YES" if dall_e_prompt else "NO"
    print(f"DALL-E Flag: {dall_e}")  # Debugging: Print the DALL-E flag

    return image_creation, dall_e_prompt, dall_e




# Function to scrape links from the page
def scrape_links(driver):
    try:
        wait = WebDriverWait(driver, 30)
        button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "rounded-xl")))
        ActionChains(driver).move_to_element(button).perform()
        driver.implicitly_wait(10)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "w-48")))

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        all_links = [a['href'] for a in soup.find_all('a', href=True) if a['href'].startswith("https://")]

        creator_website, creator_linkedin, creator_github, creator_x = "N/A", "N/A", "N/A", "N/A"
        for link in all_links:
            if 'linkedin.com/in' in link:
                creator_linkedin = link
            elif 'github.com' in link:
                creator_github = link
            elif 'x.com' in link or 'twitter.com' in link:
                creator_x = link
            elif creator_website == "N/A":
                creator_website = link

        return creator_website, creator_linkedin, creator_github, creator_x

    except Exception as e:
        print(f"Error occurred: {e}")
        return "", "", "", ""




# Function to scrape data from a single URL
def scrape_data(driver, url):
    driver.get(url)
    time.sleep(5)  # Give time for page to load
    soup = BeautifulSoup(driver.page_source, 'html.parser')

    # Extract title, description, and creator info
    title = soup.find('title').text.replace("ChatGPT - ", "").strip() if soup.find('title') else 'N/A'

    description_div = soup.find('div', class_='max-w-md text-center text-sm font-normal text-token-text-primary')
    description = description_div.get_text(strip=True) if description_div else 'N/A'

    creator_name = soup.find('div', class_='text-sm text-token-text-tertiary')
    creator_name = creator_name.get_text().replace("By ", "").strip() if creator_name else "N/A"


    creator_website, creator_linkedin, creator_github, creator_x = scrape_links(driver)


    image_element = soup.find('img', class_='h-full w-full bg-token-main-surface-secondary')


    if image_element and 'src' in image_element.attrs:
        images = [image_element['src']]
        images_str = images[0]
    else:
        images = []
        images_str = ""


    if images:
        image_creation, dall_e_prompt, dall_e = extract_image_data(images)
    else:
        image_creation, dall_e_prompt, dall_e = "", "", "NO"

    from lxml import html


    page_tree = html.fromstring(soup.prettify())

    # Find all button elements inside the target div based on their common class
    button_containers = page_tree.xpath(
        '//div[contains(@class, "flex") and contains(@class, "max-w-3xl") and contains(@class, "flex-wrap")]/button')



    button_texts = []
    for button in button_containers:

        inner_div_text = button.xpath('.//div[contains(@class, "line-clamp-3")]/text()')


        if inner_div_text:
            button_texts.append(inner_div_text[0].strip())
        else:
            button_texts.append("N/A")



    sample_msg_1 = sample_msg_2 = sample_msg_3 = prompt = "N/A"


    if len(button_texts) > 0:
        sample_msg_1 = button_texts[0]
    if len(button_texts) > 1:
        sample_msg_2 = button_texts[1]
    if len(button_texts) > 2:
        sample_msg_3 = button_texts[2]
    if len(button_texts) > 3:
        prompt = button_texts[3]



    # Return the extracted data in a dictionary
    return {
        'Retrieved Date': time.strftime("%Y-%m-%d"),
        'Retrieved Time': time.strftime("%H:%M:%S"),
        'URL': url,
        'Images': images_str,
        'DALL-E': dall_e,
        'Image Creation': image_creation,
        'DALL-E Prompt': dall_e_prompt,
        'Title': title,
        'Creator': creator_name,
        'Creator Website': creator_website,
        'Creator LinkedIn': creator_linkedin,
        'Creator GitHub': creator_github,
        'Creator X': creator_x,
        'Description': description,
        'Sample Message 1': sample_msg_1,
        'Sample Message 2': sample_msg_2,
        'Sample Message 3': sample_msg_3,
        'Prompt': prompt
    }



# Function to save data to Excel
def save_to_excel(sheet, row_num, data):
    for col_num, (key, value) in enumerate(data.items(), start=1):
        sheet.cell(row=row_num, column=col_num).value = value




# Main function to execute the script
def main():

    file_path = select_excel_file()
    if not file_path:
        return

    urls, workbook = load_urls_from_excel(file_path)
    if not urls:
        messagebox.showerror("Error", "No valid URLs found in the Excel file.")
        return

    driver = initialize_webdriver()
    login(driver)

    new_sheet = create_new_sheet(workbook)

    for idx, (url, row_num) in enumerate(urls, start=2):
        print(f"Scraping data from: {url}")
        scraped_data = scrape_data(driver, url)
        save_to_excel(new_sheet, row_num, scraped_data)
        time.sleep(random.randint(3, 6))

    workbook.save(file_path)
    time.sleep(5)
    driver.quit()
    messagebox.showinfo("Completed", "Data scraping and saving completed successfully!")


# Run the script
if __name__ == "__main__":
    main()

